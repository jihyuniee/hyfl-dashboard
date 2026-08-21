import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO
import json
import re
import gspread
import streamlit.components.v1 as components
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

st.set_page_config(page_title="한영외고 야자 대시보드", page_icon="🏫", layout="wide",
                   initial_sidebar_state="collapsed")

st.markdown("""
<style>
    html, body, [class*="css"] { font-family: 'Apple SD Gothic Neo', 'Malgun Gothic', sans-serif; }
    [data-testid="metric-container"] {
        background: white; border: 1px solid #e5e7eb;
        border-radius: 16px; padding: 16px 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }
    [data-testid="metric-container"] label { color: #6b7280; font-size: 13px; }
    [data-testid="metric-container"] [data-testid="stMetricValue"] {
        font-size: 2rem; font-weight: 700; color: #1d3a6e;
    }
    .section-title {
        font-size: 1.15rem; font-weight: 700; color: #1d3a6e;
        margin: 20px 0 10px; padding-bottom: 6px;
        border-bottom: 2px solid #2d7ef7;
    }
    .top3-card {
        background: linear-gradient(135deg, #eff4ff, #f5f0ff);
        border: 1px solid #dde8ff; border-radius: 16px;
        padding: 14px 18px; margin-bottom: 10px;
    }
    .gold   { border-left: 4px solid #f59e0b; }
    .silver { border-left: 4px solid #94a3b8; }
    .bronze { border-left: 4px solid #b45309; }
    .empty-state {
        text-align: center; padding: 32px 20px;
        color: #9ca3af; background: #f9fafb;
        border-radius: 12px; margin: 8px 0; font-size: 0.9rem;
    }
    .seat-board { overflow:auto; padding:12px; background:#f8fafc; border:1px solid #e2e8f0; border-radius:16px; }
    .seat-grid { display:grid; gap:3px; min-width:980px; }
    .seat { min-width:42px; min-height:31px; padding:3px 4px; border-radius:5px; text-decoration:none !important;
            display:flex; flex-direction:column; justify-content:center; align-items:center; line-height:1.05;
            font-size:10px; font-weight:700; box-sizing:border-box; color:#334155 !important; }
    .seat small { font-size:8px; font-weight:600; opacity:.82; margin-top:2px; }
    .seat-green { background:#dcfce7; border:2px solid #22c55e; color:#166534 !important; }
    .seat-red { background:#fee2e2; border:2px solid #ef4444; color:#991b1b !important; }
    .seat-orange { background:#ffedd5; border:2px solid #f97316; color:#9a3412 !important; }
    .seat-neutral { background:#f1f5f9; border:1px solid #cbd5e1; color:#64748b !important; }
    .seat-free { background:#fff; border:3px solid #ef4444; }
    .seat-free.seat-green { background:#dcfce7; border-color:#ef4444; box-shadow:inset 0 0 0 2px #22c55e; }
    .seat-free.seat-orange { background:#ffedd5; border-color:#ef4444; box-shadow:inset 0 0 0 2px #f97316; }
    .seat-unavailable { color:#94a3b8 !important; border:1px solid #94a3b8;
        background:repeating-linear-gradient(135deg,#e2e8f0,#e2e8f0 5px,#f8fafc 5px,#f8fafc 10px); }
    .seat-selected { outline:3px solid #2563eb; outline-offset:2px; }
    .legend { display:flex; gap:12px; flex-wrap:wrap; align-items:center; font-size:12px; color:#475569; margin:8px 0 14px; }
    .legend-dot { width:13px; height:13px; border-radius:4px; display:inline-block; margin-right:4px; vertical-align:-2px; }
    #MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ── 유틸 ─────────────────────────────────────────────
def empty_state(msg="데이터가 없습니다."):
    st.markdown(f'<div class="empty-state">📭 {msg}</div>', unsafe_allow_html=True)

def safe_int(v):
    try: return int(float(v))
    except: return None

def is_valid(v):
    return v is not None and str(v) not in ['미지정', '', 'nan', 'None']

def get_week_range(offset=0):
    today = now_kst().date()
    mon   = today - timedelta(days=today.weekday()) + timedelta(weeks=offset)
    return mon, mon + timedelta(days=6)

def make_label(g, k):
    gi, ki = safe_int(g), safe_int(k)
    if gi and ki: return f"{gi}-{ki}반"
    return "미확인"

MEDALS = ['🥇','🥈','🥉']
SCOPES = ["https://www.googleapis.com/auth/spreadsheets",
          "https://www.googleapis.com/auth/drive.readonly"]

KST = ZoneInfo("Asia/Seoul")

def now_kst():
    return datetime.now(KST)

def normalize_student_id(row):
    if is_valid(row.get('학번')):
        return str(safe_int(row.get('학번')))
    g, k, n = safe_int(row.get('학년')), safe_int(row.get('반')), safe_int(row.get('번호'))
    return str(g * 10000 + k * 100 + n) if g and k and n else ''

def normalize_seat(v):
    return str(v).strip().upper().replace(' ', '') if is_valid(v) else ''

def find_seat_column(columns):
    """출석 시트의 좌석/좌석번호/선택 좌석 등 다양한 열 제목을 찾는다."""
    names = [str(c).strip() for c in columns]
    if '좌석' in names:
        return '좌석'
    for name in names:
        compact = re.sub(r'[\s_\-()]', '', name).lower()
        if '좌석' in compact or 'seat' in compact:
            return name
    return None

@st.cache_data(ttl=60)
def load_optional_sheet(key, ws_name):
    try:
        creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
        ws = gspread.authorize(creds).open_by_key(key).worksheet(ws_name)
        records = ws.get_all_records()
        if not records:
            return pd.DataFrame()
        df = pd.DataFrame(records)
        df.columns = df.columns.astype(str).str.strip()
        return df
    except gspread.WorksheetNotFound:
        return pd.DataFrame()
    except Exception:
        return pd.DataFrame()

def border_is_thick(cell):
    return any(side.style in ('medium','thick') for side in
               (cell.border.left, cell.border.right, cell.border.top, cell.border.bottom))

def parse_setup_files(participation_file, seating_file):
    """원본 엑셀을 비공개 Google Sheet에 저장할 정규화 표로 변환한다."""
    pw = load_workbook(BytesIO(participation_file.getvalue()), data_only=True)
    student_map = {}
    for grade_name in ('1학년','2학년','3학년'):
        if grade_name not in pw.sheetnames:
            continue
        ws = pw[grade_name]
        grade = int(grade_name[0])
        for r in range(4, ws.max_row + 1):
            sid, name = ws.cell(r,2).value, ws.cell(r,3).value
            if not isinstance(sid, (int,float)) or not name:
                continue
            sid = str(int(sid))
            item = {'학번':sid, '성명':str(name).strip(), '학년':grade,
                    '반':safe_int(sid[1:3]), '번호':safe_int(sid[3:5])}
            for c, day in zip(range(4,9), '월화수목금'):
                item[day] = 1 if ws.cell(r,c).value not in (None,'',0) else 0
            # 원본에 같은 학번이 여러 번 있으면 마지막(최종 수정) 행을 사용한다.
            student_map[sid] = item

    sw = load_workbook(BytesIO(seating_file.getvalue()), data_only=True)
    seat_pattern = re.compile(r'^[A-Z]-\d+$')
    seats, seen = [], set()
    for ws in sw.worksheets:
        floor = '1학년 2층' if '2층' in ws.title else '2·3학년 4층'
        for row in ws.iter_rows():
            for cell in row:
                raw = cell.value
                if not isinstance(raw, str) or not seat_pattern.match(raw.strip()):
                    continue
                seat = raw.strip().upper()
                right = ws.cell(cell.row, cell.column + 1)
                sid = str(int(right.value)) if isinstance(right.value,(int,float)) and 10000 <= right.value < 40000 else ''
                diagonal = bool(cell.border.diagonalUp or cell.border.diagonalDown or
                                right.border.diagonalUp or right.border.diagonalDown)
                thick = border_is_thick(cell) or border_is_thick(right)
                if diagonal: seat_type = '사용불가'
                elif sid: seat_type = '지정석'
                elif thick: seat_type = '자유석'
                else: continue
                key = (floor, seat)
                if key in seen: continue
                seen.add(key)
                prefix = seat.split('-')[0]
                room = ('내실 E' if prefix == 'E' else '외실 G·H') if '4층' in floor else f'1학년 {prefix}구역'
                seats.append({'층':floor, '공간':room, '좌석':seat, '행':cell.row, '열':cell.column,
                              '좌석유형':seat_type, '지정학번':sid})
    return pd.DataFrame(student_map.values()), pd.DataFrame(seats)

def replace_worksheet(key, title, df):
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
    book = gspread.authorize(creds).open_by_key(key)
    try:
        ws = book.worksheet(title); ws.clear()
    except gspread.WorksheetNotFound:
        ws = book.add_worksheet(title=title, rows=max(100, len(df)+10), cols=max(12, len(df.columns)+2))
    values = [df.columns.tolist()] + df.fillna('').astype(str).values.tolist()
    ws.update(values=values, range_name='A1')

def setup_data_ui(key):
    with st.expander('⚙️ 좌석판 초기 데이터 설정', expanded=False):
        st.caption('원본 엑셀은 GitHub에 저장되지 않고 비공개 Google Sheet에 정규화하여 저장됩니다.')
        configured_pin = str(st.secrets.get('supervisor_pin', '')).strip()
        if not configured_pin:
            st.info('Streamlit secrets에 supervisor_pin을 설정하면 이곳에서 원본 엑셀을 등록할 수 있습니다.')
            return
        entered = st.text_input('감독교사 PIN', type='password', key='setup_pin')
        if entered != configured_pin: return
        p_file = st.file_uploader('야간 자기주도학습 참여조사', type=['xlsx'], key='participation_upload')
        s_file = st.file_uploader('자기주도학습 좌석배치표', type=['xlsx'], key='seating_upload')
        if st.button('좌석판 데이터 등록', disabled=not (p_file and s_file), type='primary'):
            try:
                students, seats = parse_setup_files(p_file, s_file)
                replace_worksheet(key, '야자신청', students)
                replace_worksheet(key, '좌석설정', seats)
                st.cache_data.clear()
                st.success(f'등록 완료: 학생 {len(students)}명 · 좌석 {len(seats)}석')
                st.rerun()
            except Exception as e:
                st.error(f'등록 실패: {e}')

@st.cache_data(ttl=300)
def load_data(key, ws_name):
    try:
        creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
        ws    = gspread.authorize(creds).open_by_key(key).worksheet(ws_name)
        data  = ws.get_all_records()
        if not data: return pd.DataFrame()
        df = pd.DataFrame(data)
        df.columns = df.columns.str.strip()
        seat_col = find_seat_column(df.columns)
        if seat_col and seat_col != '좌석':
            df['좌석'] = df[seat_col]
        if '날짜' in df.columns:
            df['날짜'] = pd.to_datetime(df['날짜'], errors='coerce')
            df = df.dropna(subset=['날짜'])
            df['날짜'] = df['날짜'].dt.normalize()
            df['요일'] = df['날짜'].dt.dayofweek.map({0:'월',1:'화',2:'수',3:'목',4:'금',5:'토',6:'일'})
        for col in ['학년','반','번호']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        # ★ 핵심: '2~3교시' → '3교시' 로 통일
        if '교시' in df.columns:
            df['교시'] = df['교시'].str.strip().replace('2~3교시', '3교시')
        def get_dept(k):
            try:
                n = int(k)
                if n<=2:  return '중국어과'
                if n<=4:  return '일본어과'
                if n<=6:  return '독일어과'
                if n<=8:  return '프랑스어과'
                if n<=10: return '스페인어과'
            except: pass
            return '기타'
        if '반' in df.columns:
            df['어학과'] = df['반'].apply(get_dept)
        return df
    except Exception as e:
        st.error(f"데이터 로드 오류: {e}")
        return pd.DataFrame()

@st.cache_data(ttl=300)
def load_student_list(key):
    """학생목록 시트 로드 — 같은 스프레드시트 내 '학생목록' 워크시트"""
    try:
        creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=SCOPES)
        ws    = gspread.authorize(creds).open_by_key(key).worksheet("학생목록")
        data  = ws.get_all_records()
        if not data: return pd.DataFrame()
        df = pd.DataFrame(data)
        df.columns = df.columns.str.strip()
        for col in ['학년','반','번호']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        return df
    except Exception as e:
        st.warning(f"학생목록 시트 로드 실패: {e}")
        return pd.DataFrame()

def filter_valid(df):
    if df.empty: return df
    d = df.copy()
    for col in ['학년','반']:
        if col in d.columns:
            d = d[d[col].notna()]
            d = d[d[col].astype(str).str.strip().isin(['미지정','']) == False]
    return d

def filter_period(df, start, end):
    if df.empty or '날짜' not in df.columns: return df
    return df[(df['날짜']>=pd.Timestamp(start)) & (df['날짜']<=pd.Timestamp(end))].copy()

# ── 기간 필터 컴포넌트 (TAB 2~5 공용) ────────────────
def period_filter_ui(tab_key):
    s_key = f"{tab_key}_start"
    e_key = f"{tab_key}_end"
    today = now_kst().date()
    if s_key not in st.session_state:
        st.session_state[s_key] = today - timedelta(days=today.weekday())
    if e_key not in st.session_state:
        st.session_state[e_key] = today

    st.markdown("""<div style='background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;
    padding:10px 16px;margin-bottom:12px'>
    <span style='font-size:13px;color:#64748b;font-weight:600'>📅 기간 선택</span></div>""",
    unsafe_allow_html=True)

    c1,c2,c3,c4,c5,c6 = st.columns([1,1,1,1,2,1])
    rerun_needed = False
    with c1:
        if st.button("이번 주", key=f"{tab_key}_w0", use_container_width=True):
            s,e = get_week_range(0); st.session_state[s_key]=s; st.session_state[e_key]=e; rerun_needed=True
    with c2:
        if st.button("지난 주", key=f"{tab_key}_w1", use_container_width=True):
            s,e = get_week_range(-1); st.session_state[s_key]=s; st.session_state[e_key]=e; rerun_needed=True
    with c3:
        if st.button("이번 달", key=f"{tab_key}_m0", use_container_width=True):
            st.session_state[s_key]=today.replace(day=1); st.session_state[e_key]=today; rerun_needed=True
    with c4:
        if st.button("전체", key=f"{tab_key}_all", use_container_width=True):
            st.session_state[s_key]=today.replace(year=today.year-1); st.session_state[e_key]=today; rerun_needed=True
    with c5:
        dr = st.date_input("기간", value=(st.session_state[s_key], st.session_state[e_key]),
                           format="YYYY/MM/DD", label_visibility="collapsed", key=f"{tab_key}_dr")
        if isinstance(dr,(list,tuple)) and len(dr)==2:
            st.session_state[s_key]=dr[0]; st.session_state[e_key]=dr[1]
    with c6:
        if st.button("🔄", key=f"{tab_key}_ref", use_container_width=True, help="새로고침"):
            st.cache_data.clear(); rerun_needed=True
    if rerun_needed: st.rerun()
    return st.session_state[s_key], st.session_state[e_key]

# ── 고정값 ────────────────────────────────────────────
sheet_key = "1LH_AI8jvW-vNn9I8wsj8lIot16vuLzqyjbZfDqcNgM8"
ws_name   = "출석기록"
today     = now_kst().date()

st.markdown("---")

# ── 데이터 로드 ───────────────────────────────────────
with st.spinner("데이터 불러오는 중..."):
    df_all      = load_data(sheet_key, ws_name)
    df_students = load_student_list(sheet_key)
    df_applications = load_optional_sheet(sheet_key, '야자신청')
    df_seats = load_optional_sheet(sheet_key, '좌석설정')

if df_all.empty:
    st.markdown("""<div style='text-align:center;padding:60px 20px;color:#6b7280'>
        <div style='font-size:40px;margin-bottom:12px'>📭</div>
        <h3>데이터가 없습니다</h3></div>""", unsafe_allow_html=True)
    st.stop()

df_valid = filter_valid(df_all)

# ── 대시보드 헤더 ─────────────────────────────────────
st.markdown(f"""
<div style='display:flex;align-items:center;gap:12px;margin-bottom:4px'>
  <span style='font-size:2rem'>🏫</span>
  <div>
    <h1 style='margin:0;font-size:1.7rem;color:#1d3a6e'>야간자율학습 출석 대시보드</h1>
    <p style='margin:0;color:#6b7280;font-size:0.82rem'>한영외국어고등학교 &nbsp;·&nbsp; 전체 {len(df_valid)}건</p>
  </div>
</div>""", unsafe_allow_html=True)

tab_seat, tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "🪑 감독 좌석판", "🏠 오늘 현황", "🏆 TOP6 시상", "📈 주차별 추이",
    "🏫 학년·반별", "🌍 어학과별", "👩‍🏫 담임용 조회"
])

# ══════════════════════════════════════════════════
# 감독 좌석판: 신청 요일 + 지정석 + 실제 QR 좌석 비교
# ══════════════════════════════════════════════════
with tab_seat:
    st.markdown("<div class='section-title'>🪑 실시간 야자 감독 좌석판</div>", unsafe_allow_html=True)
    st.caption('iPad 좌석판 v5 · 빈 공간 없이 자동 맞춤 · 두 손가락 확대 · 좌석을 누르면 상세 정보가 표시됩니다.')
    setup_data_ui(sheet_key)

    if df_applications.empty or df_seats.empty:
        empty_state("좌석판 초기 데이터가 없습니다. 감독교사 PIN으로 참여조사와 좌석배치표를 한 번 등록해 주세요.")
    else:
        seat_day = st.date_input('조회 날짜', value=today, max_value=today,
                                 format='YYYY/MM/DD', key='seat_board_day')
        period_values = df_valid.get('교시', pd.Series(dtype=str)).astype(str).unique().tolist()
        period_options = [p for p in ['1교시','3교시'] if p in period_values] or ['1교시','3교시']
        c_period, c_floor, c_room, c_refresh = st.columns([1,1.4,1.5,.5])
        with c_period:
            seat_period = st.selectbox('교시', period_options, key='seat_board_period')
        with c_floor:
            floors = df_seats['층'].dropna().astype(str).unique().tolist()
            seat_floor = st.selectbox('자습실', floors, key='seat_board_floor')
        floor_df = df_seats[df_seats['층'].astype(str) == seat_floor].copy()
        with c_room:
            rooms = floor_df['공간'].dropna().astype(str).unique().tolist()
            room_choice = st.selectbox('구역', ['전체'] + rooms, key='seat_board_room')
        with c_refresh:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            if st.button('🔄', help='새로고침', key='seat_board_refresh'):
                st.cache_data.clear(); st.rerun()

        day_name = '월화수목금토일'[seat_day.weekday()]
        apps = df_applications.copy()
        apps['학번키'] = apps['학번'].apply(lambda x: str(safe_int(x)) if safe_int(x) else '')
        apps_by_id = apps.set_index('학번키').to_dict('index')
        expected_ids = set(apps.loc[pd.to_numeric(apps[day_name], errors='coerce').fillna(0) > 0, '학번키']) \
                       if day_name in apps.columns else set()

        checkins = df_valid.copy()
        checkins = checkins[checkins['날짜'] == pd.Timestamp(seat_day)] if '날짜' in checkins.columns else checkins.iloc[0:0]
        if '교시' in checkins.columns:
            checkins = checkins[checkins['교시'].astype(str) == seat_period]
        checkins['학번키'] = checkins.apply(normalize_student_id, axis=1)
        seat_col = find_seat_column(checkins.columns)
        seat_source = checkins[seat_col] if seat_col else pd.Series('', index=checkins.index)
        if seat_col:
            checkins['좌석'] = seat_source
        checkins['좌석키'] = seat_source.apply(normalize_seat)
        checked_by_id = {r['학번키']:r for _,r in checkins.iterrows() if r['학번키']}
        checked_by_seat = {r['좌석키']:r for _,r in checkins.iterrows() if r['좌석키']}

        fixed = df_seats[df_seats['좌석유형'].astype(str) == '지정석'].copy()
        assigned_by_id = {str(safe_int(r['지정학번'])):normalize_seat(r['좌석'])
                          for _,r in fixed.iterrows() if safe_int(r['지정학번'])}
        mismatched_ids = {sid for sid,r in checked_by_id.items()
                          if sid in assigned_by_id and normalize_seat(r.get('좌석')) != assigned_by_id[sid]}
        expected_fixed = expected_ids.intersection(assigned_by_id)
        normal_count = sum(sid in checked_by_id and sid not in mismatched_ids for sid in expected_fixed)
        missing_count = sum(sid not in checked_by_id for sid in expected_fixed)

        m1,m2,m3,m4 = st.columns(4)
        m1.metric('오늘 지정석 대상', f'{len(expected_fixed)}명')
        m2.metric('정상 체크인', f'{normal_count}명')
        m3.metric('미체크인', f'{missing_count}명')
        m4.metric('지정석 불일치', f'{len(mismatched_ids)}명')

        st.markdown("""<div class='legend'>
          <span><i class='legend-dot' style='background:#dcfce7;border:2px solid #22c55e'></i>정상 체크인</span>
          <span><i class='legend-dot' style='background:#fee2e2;border:2px solid #ef4444'></i>오늘 신청·미체크인</span>
          <span><i class='legend-dot' style='background:#ffedd5;border:2px solid #f97316'></i>지정석 불일치</span>
          <span><i class='legend-dot' style='background:white;border:3px solid #ef4444'></i>자유석</span>
          <span><i class='legend-dot' style='background:#e2e8f0;border:1px solid #94a3b8'></i>사용 불가</span>
        </div>""", unsafe_allow_html=True)

        board_df = floor_df if room_choice == '전체' else floor_df[floor_df['공간'].astype(str) == room_choice]
        board_df = board_df.copy()
        board_df['행'] = pd.to_numeric(board_df['행'], errors='coerce')
        board_df['열'] = pd.to_numeric(board_df['열'], errors='coerce')
        board_df = board_df.dropna(subset=['행','열'])
        if board_df.empty:
            empty_state('표시할 좌석이 없습니다.')
        else:
            # 엑셀의 빈 행 높이를 그대로 늘이지 않고, 실제 좌석 행만 촘촘하게 재배치한다.
            unique_rows = sorted(int(v) for v in board_df['행'].unique())
            row_y, current_y, previous_row = {}, 14, None
            for excel_row in unique_rows:
                if previous_row is not None:
                    gap = excel_row - previous_row
                    current_y += 32 if gap == 1 else (42 if gap <= 3 else 54)
                row_y[excel_row] = current_y
                previous_row = excel_row
            min_col = int(board_df['열'].min())
            max_col = int(board_df['열'].max())
            board_width = (max_col - min_col + 1) * 27 + 70
            board_height = max(row_y.values()) + 55

            def map_excel_y(excel_row):
                """좌석이 없는 행에 있는 시설물도 압축된 좌석판 높이에 맞춰 배치한다."""
                if excel_row in row_y:
                    return row_y[excel_row]
                lower = [r for r in unique_rows if r < excel_row]
                upper = [r for r in unique_rows if r > excel_row]
                if lower and upper:
                    lo, hi = max(lower), min(upper)
                    ratio = (excel_row - lo) / (hi - lo)
                    return round(row_y[lo] + (row_y[hi] - row_y[lo]) * ratio)
                if lower:
                    lo = max(lower)
                    return row_y[lo] + min(54, (excel_row - lo) * 12)
                hi = min(upper)
                return max(8, row_y[hi] - min(54, (hi - excel_row) * 12))

            seat_items = []
            for _, seat_row in board_df.iterrows():
                seat = normalize_seat(seat_row['좌석'])
                seat_type = str(seat_row['좌석유형'])
                sid = str(safe_int(seat_row.get('지정학번'))) if safe_int(seat_row.get('지정학번')) else ''
                actual = checked_by_seat.get(seat)
                status, status_label = 'neutral', '오늘 미신청'
                display_sid = sid
                if seat_type == '사용불가':
                    status, status_label, display_sid = 'unavailable', '사용 불가', ''
                elif seat_type == '자유석':
                    if actual is not None:
                        actual_sid = normalize_student_id(actual)
                        status = 'free-mismatch' if actual_sid in mismatched_ids else 'free-checked'
                        status_label, display_sid = ('지정석 불일치' if actual_sid in mismatched_ids else '자유석 체크인'), actual_sid
                    else:
                        status, status_label, display_sid = 'free', '비어 있는 자유석', ''
                elif actual is not None:
                    actual_sid = normalize_student_id(actual)
                    status = 'mismatch' if actual_sid in mismatched_ids else 'checked'
                    status_label, display_sid = ('지정석 불일치' if actual_sid in mismatched_ids else '정상 체크인'), actual_sid
                elif sid in mismatched_ids:
                    status, status_label = 'mismatch', f"다른 좌석({normalize_seat(checked_by_id[sid].get('좌석'))})에서 체크인"
                elif sid in expected_ids:
                    status, status_label = 'missing', '오늘 신청 · 미체크인'
                info = apps_by_id.get(display_sid, {})
                days = ' · '.join(d for d in '월화수목금' if safe_int(info.get(d)) == 1)
                # 지정석이 비어 있어도 그 학생이 다른 좌석에서 체크인했다면
                # 학번 기준 기록에서 실제 좌석과 시각을 가져온다.
                student_checkin = actual if actual is not None else checked_by_id.get(display_sid)
                checkin_time = str(student_checkin.get('시간','')) if student_checkin is not None else ''
                seat_items.append({
                    'seat':seat, 'studentId':display_sid, 'name':str(info.get('성명','')),
                    'grade':str(info.get('학년','')), 'classNo':str(info.get('반','')), 'number':str(info.get('번호','')),
                    'days':days, 'assignedSeat':assigned_by_id.get(display_sid, '자유석' if seat_type == '자유석' else seat),
                    'actualSeat':normalize_seat(student_checkin.get('좌석')) if student_checkin is not None else '',
                    'checkinTime':checkin_time, 'status':status, 'statusLabel':status_label,
                    'seatType':seat_type, 'x':(int(seat_row['열']) - min_col) * 27 + 12,
                    'y':row_y[int(seat_row['행'])]
                })

            # 전체 보기에서는 엑셀의 큰 여백을 그대로 두지 않고 패드의 가로 화면에 맞게
            # 각 자습실의 위쪽을 맞춘 뒤 시설 통로를 사이에 둔다.
            facility_x_override = None
            if room_choice == '전체' and '2층' in seat_floor:
                left_seats = [s for s in seat_items if s['seat'].startswith('C-')]
                right_seats = [s for s in seat_items if s['seat'].split('-')[0] in {'A','B'}]
                if left_seats and right_seats:
                    left_x0, left_y0 = min(s['x'] for s in left_seats), min(s['y'] for s in left_seats)
                    right_x0, right_y0 = min(s['x'] for s in right_seats), min(s['y'] for s in right_seats)
                    left_width = max(s['x'] + 54 for s in left_seats) - left_x0
                    corridor_width = 176
                    for s in left_seats:
                        s['x'] = s['x'] - left_x0 + 30
                        s['y'] = s['y'] - left_y0 + 30
                    for s in right_seats:
                        s['x'] = s['x'] - right_x0 + 30 + left_width + corridor_width
                        s['y'] = s['y'] - right_y0 + 30
                    facility_x_override = 30 + left_width + 36

            # 4층 내실(E)도 외실(G·H) 왼쪽에 나란히 놓아 상단 공백을 없앤다.
            outer_shift_x = 0
            if room_choice == '전체' and '4층' in seat_floor:
                inner_seats = [s for s in seat_items if s['seat'].startswith('E-')]
                outer_seats = [s for s in seat_items if s['seat'].split('-')[0] in {'G','H'}]
                if inner_seats and outer_seats:
                    inner_left = min(s['x'] for s in inner_seats)
                    inner_top = min(s['y'] for s in inner_seats)
                    inner_right = max(s['x'] + 54 for s in inner_seats)
                    outer_left = min(s['x'] for s in outer_seats)
                    outer_top = min(s['y'] for s in outer_seats)
                    inner_width = inner_right - inner_left
                    outer_shift_x = inner_width + 58
                    for s in inner_seats:
                        s['x'] = s['x'] - inner_left + 30
                        s['y'] = s['y'] - inner_top + outer_top
                    for s in outer_seats:
                        s['x'] = s['x'] - outer_left + 30 + outer_shift_x

            def seat_bounds(prefixes, padding=12):
                selected = [s for s in seat_items if s['seat'].split('-')[0] in prefixes]
                if not selected:
                    return None
                left = min(s['x'] for s in selected) - padding
                top = min(s['y'] for s in selected) - padding
                right = max(s['x'] + 54 for s in selected) + padding
                bottom = max(s['y'] + 30 for s in selected) + padding
                return {'x':left, 'y':top, 'w':right-left, 'h':bottom-top}

            # 원본 좌석배치표에 표시된 공간과 시설물을 같은 위치 관계로 복원한다.
            map_items = []
            if room_choice == '전체' and '2층' in seat_floor:
                left_zone, right_zone = seat_bounds({'C'}, 18), seat_bounds({'A','B'}, 18)
                if left_zone:
                    map_items.append({**left_zone, 'kind':'zone zone-left', 'label':'좌측 자습실', 'sub':'94석'})
                if right_zone:
                    map_items.append({**right_zone, 'kind':'zone zone-right', 'label':'우측 자습실', 'sub':'143석'})
                facility_x = facility_x_override if facility_x_override is not None else (23 - min_col) * 27 + 12
                map_items.extend([
                    {'x':facility_x, 'y':map_excel_y(10), 'w':54, 'h':max(72, map_excel_y(18)-map_excel_y(10)+30),
                     'kind':'facility vertical', 'label':'화장실', 'sub':''},
                    {'x':facility_x, 'y':map_excel_y(22), 'w':108, 'h':max(92, map_excel_y(29)-map_excel_y(22)+30),
                     'kind':'facility', 'label':'충전', 'sub':'스테이션'},
                    {'x':facility_x, 'y':map_excel_y(34), 'w':54, 'h':max(72, map_excel_y(38)-map_excel_y(34)+30),
                     'kind':'facility vertical', 'label':'감독실', 'sub':''},
                ])
                board_height = max(board_height, *(i['y'] + i['h'] + 30 for i in map_items))
            elif room_choice == '전체' and '4층' in seat_floor:
                inner_zone, outer_zone = seat_bounds({'E'}, 18), seat_bounds({'G','H'}, 18)
                if inner_zone:
                    map_items.append({**inner_zone, 'kind':'zone zone-inner', 'label':'내실', 'sub':'40석'})
                if outer_zone:
                    map_items.append({**outer_zone, 'kind':'zone zone-outer', 'label':'외실', 'sub':'270석'})
                supervisor_x = (36 - min_col) * 27 + 12 - outer_left + 30 + outer_shift_x
                map_items.append({'x':supervisor_x, 'y':map_excel_y(24), 'w':108, 'h':58,
                                  'kind':'facility supervisor', 'label':'감독석', 'sub':''})
                board_width = max(board_width, supervisor_x + 132)
                board_height = max(board_height, *(i['y'] + i['h'] + 30 for i in map_items))

            # 실제 좌석·시설물의 최상단/최좌측을 기준으로 내부 여백까지 잘라낸다.
            # 원본 엑셀의 빈 행 좌표가 남아 좌석판 위가 비는 현상을 방지한다.
            content_left = min(
                min(s['x'] for s in seat_items),
                min((i['x'] for i in map_items), default=float('inf'))
            )
            content_top = min(
                min(s['y'] for s in seat_items),
                min((i['y'] for i in map_items), default=float('inf'))
            )
            shift_x, shift_y = content_left - 18, content_top - 18
            for item in seat_items:
                item['x'] -= shift_x
                item['y'] -= shift_y
            for item in map_items:
                item['x'] -= shift_x
                item['y'] -= shift_y

            # 실제 콘텐츠의 끝까지만 캔버스를 만들고 화면 안에 비율대로 맞춘다.
            content_right = max(
                max(s['x'] + 54 for s in seat_items),
                max((i['x'] + i['w'] for i in map_items), default=0)
            )
            content_bottom = max(
                max(s['y'] + 30 for s in seat_items),
                max((i['y'] + i['h'] for i in map_items), default=0)
            )
            board_width = max(360, content_right + 24)
            board_height = max(300, content_bottom + 36)

            payload = json.dumps(seat_items, ensure_ascii=False).replace('</', '<\\/')
            map_payload = json.dumps(map_items, ensure_ascii=False).replace('</', '<\\/')
            room_title = f'{seat_floor} · {room_choice}'
            room_title_json = json.dumps(room_title, ensure_ascii=False)
            component_html = f"""
            <!doctype html><html><head><meta charset='utf-8'><style>
              *{{box-sizing:border-box}} body{{margin:0;font-family:-apple-system,BlinkMacSystemFont,'Apple SD Gothic Neo','Malgun Gothic',sans-serif;background:white;color:#1e293b}}
              .wrap{{display:grid;grid-template-columns:minmax(0,1fr) 260px;gap:12px;height:700px;min-width:0}}
              .viewport{{position:relative;overflow:hidden;display:flex;align-items:flex-start;justify-content:center;min-width:0;min-height:0;padding:10px;background:#f8fafc;border:1px solid #dbe3ee;border-radius:14px;-webkit-overflow-scrolling:touch;touch-action:none}}
              .stage{{position:relative;flex:none}}
              .canvas{{position:absolute;left:0;top:0;width:{board_width}px;height:{board_height}px;transform-origin:top left;background:#f8fafc}}
              .seat{{position:absolute;z-index:2;width:54px;height:30px;overflow:hidden;border-radius:6px;padding:2px;background:#f1f5f9;border:1px solid #cbd5e1;color:#64748b;cursor:pointer;text-align:center;line-height:1.05}}
              .seat b{{display:block;font-size:10px;white-space:nowrap}} .seat span{{display:block;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;font-size:8px;margin-top:2px}}
              .seat:hover,.seat.active{{outline:3px solid #2563eb;outline-offset:2px;z-index:3}}
              .checked{{background:#dcfce7;border:2px solid #22c55e;color:#166534}}
              .missing{{background:#fee2e2;border:2px solid #ef4444;color:#991b1b}}
              .mismatch{{background:#ffedd5;border:2px solid #f97316;color:#9a3412}}
              .free{{background:white;border:3px solid #ef4444;color:#991b1b}}
              .free-checked{{background:#dcfce7;border:3px solid #ef4444;box-shadow:inset 0 0 0 2px #22c55e;color:#166534}}
              .free-mismatch{{background:#ffedd5;border:3px solid #ef4444;box-shadow:inset 0 0 0 2px #f97316;color:#9a3412}}
              .unavailable{{background:repeating-linear-gradient(135deg,#e2e8f0,#e2e8f0 5px,#f8fafc 5px,#f8fafc 10px);border:1px solid #94a3b8;color:#94a3b8}}
              .map-item{{position:absolute;z-index:0;pointer-events:none;display:flex;align-items:center;justify-content:center;text-align:center}}
              .facility{{z-index:1;border:2px solid #64748b;border-radius:5px;background:rgba(255,255,255,.94);font-size:12px;font-weight:800;color:#334155;line-height:1.45;box-shadow:0 2px 5px rgba(15,23,42,.06)}}
              .facility.vertical{{writing-mode:vertical-rl;letter-spacing:4px}}
              .facility.supervisor{{border-color:#475569;background:#fff7ed;color:#9a3412}}
              .zone{{border:2px solid rgba(100,116,139,.35);border-radius:12px;background:rgba(255,255,255,.24)}}
              .zone::before{{content:attr(data-label);position:absolute;left:10px;top:-12px;border:1px solid #cbd5e1;border-radius:999px;background:white;padding:3px 9px;font-size:10px;font-weight:800;color:#475569;white-space:nowrap}}
              .zone::after{{content:attr(data-sub);position:absolute;left:50%;bottom:-25px;transform:translateX(-50%);font-size:12px;font-weight:800;color:#334155;white-space:nowrap}}
              .zone-left{{background:rgba(255,237,213,.18)}} .zone-right{{background:rgba(254,249,195,.18)}}
              .zone-inner{{background:rgba(243,232,255,.22);border-color:rgba(168,85,247,.35)}} .zone-outer{{background:rgba(219,234,254,.18);border-color:rgba(59,130,246,.3)}}
              .zoom-controls{{display:flex;align-items:center;gap:6px;width:max-content;margin:0 0 14px;padding:5px;border:1px solid #cbd5e1;border-radius:10px;background:#f8fafc}}
              .zoom-controls button{{width:32px;height:32px;padding:0;border:1px solid #cbd5e1;border-radius:7px;background:white;color:#1d3a6e;font-size:18px;font-weight:800;cursor:pointer}}
              .zoom-controls button:last-child{{width:48px;font-size:11px}} .zoom-controls button:active{{background:#eff6ff}}
              .panel{{border:1px solid #dbe3ee;border-radius:14px;padding:18px;background:white;overflow:auto}}
              .panel h3{{font-size:16px;margin:0 0 4px;color:#1d3a6e}} .muted{{font-size:11px;color:#94a3b8;margin-bottom:16px}}
              .empty{{height:90%;display:flex;align-items:center;justify-content:center;text-align:center;color:#94a3b8;font-size:13px;line-height:1.6}}
              .badge{{display:inline-block;border-radius:999px;padding:5px 10px;font-size:11px;font-weight:700;margin:8px 0 15px;background:#f1f5f9}}
              .badge.checked{{background:#dcfce7;color:#166534;border:0}} .badge.missing{{background:#fee2e2;color:#991b1b;border:0}} .badge.mismatch,.badge.free-mismatch{{background:#ffedd5;color:#9a3412;border:0}}
              dl{{margin:0}} dt{{font-size:10px;color:#94a3b8;margin-top:12px}} dd{{margin:3px 0 0;font-size:13px;font-weight:650;color:#334155}}
              @media(max-width:900px){{
                .wrap{{grid-template-columns:minmax(0,1fr);grid-template-rows:minmax(0,1fr) 168px;gap:8px}}
                .panel{{padding:12px}} .empty{{height:100%}} dl{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:2px 10px}}
                dt{{margin-top:4px}} dd{{font-size:11px}}
              }}
            </style></head><body>
              <div class='wrap'>
                <div class='viewport' id='viewport'>
                  <div class='stage' id='stage'><div class='canvas' id='canvas'></div></div>
                </div>
                <aside class='panel'><div class='zoom-controls'><button id='zoomOut' aria-label='축소'>−</button><button id='zoomIn' aria-label='확대'>＋</button><button id='zoomFit'>맞춤</button></div><div id='panel'><div class='empty'>두 손가락으로 확대하거나<br>좌석을 눌러 정보를 확인하세요.</div></div></aside>
              </div>
              <script>
                const seats={payload}; const mapItems={map_payload}; const roomTitle={room_title_json};
                const naturalWidth={board_width}, naturalHeight={board_height};
                const viewport=document.getElementById('viewport'), stage=document.getElementById('stage');
                const canvas=document.getElementById('canvas'), panel=document.getElementById('panel');
                const safe=v=>String(v??'').replace(/[&<>\"']/g,m=>({{'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#39;'}}[m]));
                let fitScale=1, zoomFactor=1;
                function applyScale(){{
                  const scale=fitScale*zoomFactor;
                  canvas.style.transform=`scale(${{scale}})`;
                  stage.style.width=(naturalWidth*scale)+'px'; stage.style.height=(naturalHeight*scale)+'px';
                  const zoomed=zoomFactor>1.01;
                  viewport.style.overflow=zoomed?'auto':'hidden';
                  viewport.style.justifyContent=zoomed?'flex-start':'center';
                }}
                function fitBoard(){{
                  const styles=getComputedStyle(viewport);
                  const usableW=viewport.clientWidth-parseFloat(styles.paddingLeft)-parseFloat(styles.paddingRight);
                  const usableH=viewport.clientHeight-parseFloat(styles.paddingTop)-parseFloat(styles.paddingBottom);
                  fitScale=Math.min(usableW/naturalWidth,usableH/naturalHeight);
                  applyScale();
                }}
                document.getElementById('zoomIn').onclick=()=>{{zoomFactor=Math.min(3,zoomFactor+.25);applyScale();}};
                document.getElementById('zoomOut').onclick=()=>{{zoomFactor=Math.max(1,zoomFactor-.25);applyScale();}};
                document.getElementById('zoomFit').onclick=()=>{{zoomFactor=1;viewport.scrollTo(0,0);applyScale();}};
                let gestureStartFactor=1;
                viewport.addEventListener('gesturestart',e=>{{gestureStartFactor=zoomFactor;e.preventDefault();}},{{passive:false}});
                viewport.addEventListener('gesturechange',e=>{{zoomFactor=Math.max(1,Math.min(3,gestureStartFactor*e.scale));applyScale();e.preventDefault();}},{{passive:false}});
                const touchDistance=touches=>Math.hypot(touches[0].clientX-touches[1].clientX,touches[0].clientY-touches[1].clientY);
                let pinchDistance=0,pinchFactor=1,lastTouch=null;
                viewport.addEventListener('touchstart',e=>{{
                  if(e.touches.length===2){{pinchDistance=touchDistance(e.touches);pinchFactor=zoomFactor;lastTouch=null;e.preventDefault();}}
                  else if(e.touches.length===1 && zoomFactor>1.01) lastTouch={{x:e.touches[0].clientX,y:e.touches[0].clientY}};
                }},{{passive:false}});
                viewport.addEventListener('touchmove',e=>{{
                  if(e.touches.length===2 && pinchDistance){{
                    zoomFactor=Math.max(1,Math.min(3,pinchFactor*touchDistance(e.touches)/pinchDistance));applyScale();e.preventDefault();
                  }} else if(e.touches.length===1 && lastTouch && zoomFactor>1.01){{
                    const touch=e.touches[0];viewport.scrollBy(lastTouch.x-touch.clientX,lastTouch.y-touch.clientY);lastTouch={{x:touch.clientX,y:touch.clientY}};e.preventDefault();
                  }}
                }},{{passive:false}});
                viewport.addEventListener('touchend',e=>{{if(e.touches.length<2)pinchDistance=0;if(e.touches.length===0)lastTouch=null;}},{{passive:false}});
                mapItems.forEach(item=>{{
                  const el=document.createElement('div'); el.className='map-item '+item.kind;
                  el.style.left=item.x+'px'; el.style.top=item.y+'px'; el.style.width=item.w+'px'; el.style.height=item.h+'px';
                  el.dataset.label=item.label||''; el.dataset.sub=item.sub||'';
                  if(item.kind.includes('facility')) el.innerHTML='<div>'+safe(item.label)+(item.sub?'<br>'+safe(item.sub):'')+'</div>';
                  canvas.appendChild(el);
                }});
                seats.forEach(s=>{{
                  const el=document.createElement('button'); el.className='seat '+s.status; el.style.left=s.x+'px'; el.style.top=s.y+'px';
                  el.innerHTML='<b>'+safe(s.seat)+'</b><span>'+safe(s.studentId||s.statusLabel)+'</span>';
                  el.onclick=()=>{{document.querySelectorAll('.seat.active').forEach(x=>x.classList.remove('active'));el.classList.add('active');
                    const student=s.name ? `${{safe(s.grade)}}학년 ${{safe(s.classNo)}}반 ${{safe(s.number)}}번 ${{safe(s.name)}}` : (s.seatType==='자유석'?'현재 이용 학생 없음':'배정 학생 없음');
                    panel.innerHTML=`<h3>${{safe(s.seat)}}</h3><div class='muted'>${{safe(roomTitle)}}</div><span class='badge ${{safe(s.status)}}'>${{safe(s.statusLabel)}}</span><dl>
                    <dt>학생</dt><dd>${{student}}</dd><dt>학번</dt><dd>${{safe(s.studentId||'-')}}</dd><dt>신청 요일</dt><dd>${{safe(s.days||'-')}}</dd>
                    <dt>지정 좌석</dt><dd>${{safe(s.assignedSeat||'-')}}</dd><dt>체크인 좌석</dt><dd>${{safe(s.actualSeat||'-')}}</dd><dt>체크인 시각</dt><dd>${{safe(s.checkinTime||'-')}}</dd></dl>`;
                  }}; canvas.appendChild(el);
                }});
                new ResizeObserver(fitBoard).observe(viewport); fitBoard();
              </script></body></html>
            """
            components.html(component_html, height=715, scrolling=False)

# ══════════════════════════════════════════════════
# TAB 1: 오늘 현황
# ══════════════════════════════════════════════════
with tab1:
    if "tab1_selected_day" not in st.session_state:
        st.session_state["tab1_selected_day"] = today

    col_title_t1, col_spacer, col_cal = st.columns([3, 3, 2])
    with col_title_t1:
        st.markdown("<div style='margin-top:8px;font-size:1.1rem;font-weight:700;color:#1d3a6e'>📊 출석 현황</div>",
                    unsafe_allow_html=True)
    with col_cal:
        picked = st.date_input("", value=st.session_state["tab1_selected_day"],
                               max_value=today, format="YYYY/MM/DD",
                               label_visibility="collapsed")
        if picked != st.session_state["tab1_selected_day"]:
            st.session_state["tab1_selected_day"] = picked
            st.rerun()

    selected_day = st.session_state["tab1_selected_day"]

    dow_map = {0:'월',1:'화',2:'수',3:'목',4:'금',5:'토',6:'일'}
    dow = dow_map[selected_day.weekday()]
    is_today_flag = selected_day == today
    badge_text = f"📅 {selected_day.year}년 {selected_day.month}월 {selected_day.day}일 ({dow}){' · 오늘' if is_today_flag else ''}"

    col_badge, col_shortcuts = st.columns([3, 2])
    with col_badge:
        st.markdown(f"""<div style='margin-bottom:12px'>
          <span style='background:#eff6ff;border:1.5px solid #3b82f6;border-radius:20px;
            padding:4px 14px;font-size:12px;font-weight:600;color:#3b82f6'>{badge_text}</span>
        </div>""", unsafe_allow_html=True)
    with col_shortcuts:
        sc1, sc2 = st.columns(2)
        with sc1:
            if st.button("어제", use_container_width=True, key="tab1_yesterday"):
                st.session_state["tab1_selected_day"] = today - timedelta(days=1)
                st.rerun()
        with sc2:
            if st.button("오늘", use_container_width=True, key="tab1_today"):
                st.session_state["tab1_selected_day"] = today
                st.rerun()

    st.markdown("<hr style='border:none;border-top:2px solid #3b82f6;margin-bottom:16px'>", unsafe_allow_html=True)

    today_ts   = pd.Timestamp(selected_day)
    df_today   = df_all[df_all['날짜']==today_ts] if '날짜' in df_all.columns else pd.DataFrame()
    df_today_v = filter_valid(df_today)

    def cnt(df_t, grade=None, period=None):
        d = df_t.copy()
        if grade  and '학년' in d.columns: d = d[d['학년']==grade]
        if period and '교시' in d.columns: d = d[d['교시']==period]
        return d['이메일'].nunique() if '이메일' in d.columns and not d.empty else 0

    # ★ 교시 라벨 통일: '3교시'
    p1_label = '1교시'
    p2_label = '3교시'

    data_table = {
        '학년':   ['1학년','2학년','3학년','✅ 전체'],
        '1교시':  [cnt(df_today_v,1,p1_label),cnt(df_today_v,2,p1_label),
                   cnt(df_today_v,3,p1_label),cnt(df_today_v,None,p1_label)],
        '3교시':  [cnt(df_today_v,1,p2_label),cnt(df_today_v,2,p2_label),
                   cnt(df_today_v,3,p2_label),cnt(df_today_v,None,p2_label)],
    }
    data_table['합계'] = [a+b for a,b in zip(data_table['1교시'],data_table['3교시'])]

    tbody = ""
    for label,v1,v2,vt in zip(data_table['학년'],data_table['1교시'],data_table['3교시'],data_table['합계']):
        is_total = (label=='✅ 전체')
        bg = '#f0fdf4' if is_total else 'white'
        fw = '700'    if is_total else '500'
        fs = '1.2rem' if is_total else '1.4rem'
        tbody += f"<tr style='background:{bg}'><td style='padding:16px 24px;font-weight:{fw};font-size:1rem;border-bottom:1px solid #f3f4f6'>{label}</td><td style='padding:16px;text-align:center;font-weight:{fw};font-size:{fs};color:#3b82f6;border-bottom:1px solid #f3f4f6'>{v1}명</td><td style='padding:16px;text-align:center;font-weight:{fw};font-size:{fs};color:#8b5cf6;border-bottom:1px solid #f3f4f6'>{v2}명</td><td style='padding:16px;text-align:center;font-weight:{fw};font-size:{fs};color:#10b981;border-bottom:1px solid #f3f4f6'>{vt}명</td></tr>"

    # ★ 테이블 헤더도 '3교시'로 변경
    st.markdown(f"""<table style='width:100%;border-collapse:collapse;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(26,47,90,0.08);margin-bottom:20px'>
      <thead><tr style='background:#1d3a6e;color:white'>
        <th style='padding:16px 24px;text-align:left;font-size:0.95rem;width:22%'>학년</th>
        <th style='padding:16px;text-align:center;font-size:0.95rem;width:26%'>1교시<br><span style='font-size:0.75rem;opacity:0.8;font-weight:400'>16:10 ~ 17:40</span></th>
        <th style='padding:16px;text-align:center;font-size:0.95rem;width:26%'>3교시<br><span style='font-size:0.75rem;opacity:0.8;font-weight:400'>18:40 ~ 21:50</span></th>
        <th style='padding:16px;text-align:center;font-size:0.95rem;width:26%'>합계</th>
      </tr></thead><tbody>{tbody}</tbody></table>""", unsafe_allow_html=True)

    if df_today_v.empty:
        empty_state("선택한 날짜의 출석 데이터가 없습니다.")
    else:
        col_l, col_r = st.columns(2)
        with col_l:
            st.markdown("<div class='section-title'>학년별 참여 인원</div>", unsafe_allow_html=True)
            if '학년' in df_today_v.columns and '이메일' in df_today_v.columns:
                gd = df_today_v.groupby('학년')['이메일'].nunique().reset_index().rename(columns={'이메일':'학생수'})
                gd['학년명'] = gd['학년'].apply(lambda x: f"{int(x)}학년")
                fig = px.bar(gd, x='학년명', y='학생수', color='학생수',
                             color_continuous_scale=['#93c5fd','#1d4ed8'], text='학생수')
                fig.update_traces(textposition='outside')
                fig.update_layout(showlegend=False, coloraxis_showscale=False,
                                  plot_bgcolor='white', paper_bgcolor='white',
                                  margin=dict(t=20,b=0), height=260)
                st.plotly_chart(fig, use_container_width=True)
        with col_r:
            st.markdown("<div class='section-title'>반별 참여 인원</div>", unsafe_allow_html=True)
            if '학년' in df_today_v.columns and '반' in df_today_v.columns:
                cd = df_today_v.groupby(['학년','반'])['이메일'].nunique().reset_index().rename(columns={'이메일':'학생수'})
                cd['반명'] = cd.apply(lambda r: make_label(r['학년'],r['반']), axis=1)
                cd = cd[cd['반명']!='미확인']
                if not cd.empty:
                    fig2 = px.bar(cd, x='반명', y='학생수', color='학년',
                                  text='학생수', color_continuous_scale='Blues')
                    fig2.update_traces(textposition='outside')
                    fig2.update_layout(plot_bgcolor='white', paper_bgcolor='white',
                                       margin=dict(t=20,b=0), height=260,
                                       showlegend=False, coloraxis_showscale=False)
                    st.plotly_chart(fig2, use_container_width=True)
                else:
                    empty_state("반별 데이터 없음")

        st.markdown("<div class='section-title'>출석 명단</div>", unsafe_allow_html=True)
        show_cols = [c for c in ['학년','반','번호','이름','교시','좌석','시간'] if c in df_today_v.columns]
        if show_cols:
            sort_by = [c for c in ['학년','반','번호'] if c in show_cols]
            st.dataframe(df_today_v[show_cols].sort_values(sort_by) if sort_by else df_today_v[show_cols],
                         use_container_width=True, height=320)

# ══════════════════════════════════════════════════
# TAB 2: TOP3 시상
# ══════════════════════════════════════════════════
with tab2:
    start_date, end_date = period_filter_ui("t2")
    df = filter_period(df_valid, start_date, end_date)
    st.markdown("<div class='section-title'>🏆 기간 내 TOP6 현황</div>", unsafe_allow_html=True)
    st.caption(f"기간: {start_date} ~ {end_date}")

    if df.empty or '이메일' not in df.columns:
        empty_state("선택 기간에 데이터가 없습니다.")
    else:
        grp_cols = [c for c in ['이메일','이름','학년','반'] if c in df.columns]

        # ★ TOP6는 하루에 1교시+3교시를 모두 출석한 날만 인정한다 (한 교시만 출석한 날은 카운트하지 않음)
        if '교시' in df.columns and '날짜' in df.columns:
            day_periods = df.groupby(grp_cols + ['날짜'])['교시'].apply(lambda s: set(s)).reset_index(name='교시집합')
            day_periods['완전출석'] = day_periods['교시집합'].apply(lambda s: {'1교시','3교시'}.issubset(s))
            full_days = day_periods[day_periods['완전출석']]
            stu = full_days.groupby(grp_cols).agg(체크인수=('날짜','nunique')).reset_index()
        else:
            stu = df.groupby(grp_cols).agg(체크인수=('날짜','nunique')).reset_index()

        def show_top3(data, title, score_col='체크인수'):
            st.markdown(f"**{title}**")
            if data.empty: empty_state("데이터 없음"); return
            sorted_data = data.sort_values(score_col, ascending=False).reset_index(drop=True)
            # ★ 6위와 동점인 학생도 모두 "공동 6위"로 함께 표시
            cutoff_val = sorted_data[score_col].iloc[min(5, len(sorted_data)-1)]
            top = sorted_data[sorted_data[score_col] >= cutoff_val].reset_index(drop=True)
            for i, row in top.iterrows():
                rank  = int((top[score_col] > row[score_col]).sum()) + 1
                medal = MEDALS[rank-1] if rank<=3 else f"{rank}위"
                cls   = ['gold','silver','bronze'][rank-1] if rank<=3 else ''
                name  = row.get('이름',''); grade=safe_int(row.get('학년','')); klass=safe_int(row.get('반',''))
                dept  = row['어학과'] if '어학과' in row.index else ''
                info  = f"{grade}학년 {klass}반 {dept}" if grade and klass else ''
                cnt_v = int(row[score_col])
                st.markdown(f"""<div class="top3-card {cls}">
                  <span style="font-size:1.3rem">{medal}</span>
                  <strong style="font-size:1rem;margin-left:8px">{name}</strong>
                  <span style="color:#6b7280;font-size:0.85rem;margin-left:8px">{info}</span>
                  <span style="float:right;color:#2d7ef7;font-weight:700">{cnt_v}일 (1·3교시 모두 출석)</span>
                </div>""", unsafe_allow_html=True)

        st.markdown("<div class='section-title'>🌟 전체 학생 TOP6</div>", unsafe_allow_html=True)
        if '어학과' not in stu.columns and '반' in df.columns:
            stu['어학과'] = stu['반'].apply(lambda k: (
                '중국어과' if k<=2 else '일본어과' if k<=4 else '독일어과'
                if k<=6 else '프랑스어과' if k<=8 else '스페인어과') if pd.notna(k) else '')
        show_top3(stu, "🏅 전체 TOP6")
        st.markdown("---")

        st.markdown("<div class='section-title'>📚 학년별 TOP6</div>", unsafe_allow_html=True)
        gc1,gc2,gc3 = st.columns(3)
        for col,grade_n in zip([gc1,gc2,gc3],[1,2,3]):
            with col:
                show_top3(stu[stu['학년']==grade_n] if '학년' in stu.columns else pd.DataFrame(),
                          f"🎓 {grade_n}학년 TOP6")
        st.markdown("---")

        st.markdown("<div class='section-title'>🏫 반별 TOP6</div>", unsafe_allow_html=True)
        if '학년' in df.columns and '반' in df.columns and '이메일' in df.columns:
            if '교시' in df.columns and '날짜' in df.columns:
                cls_day_periods = df.groupby(['학년','반','이메일','날짜'])['교시'].apply(lambda s: set(s)).reset_index(name='교시집합')
                cls_day_periods['완전출석'] = cls_day_periods['교시집합'].apply(lambda s: {'1교시','3교시'}.issubset(s))
                cls_full_days = cls_day_periods[cls_day_periods['완전출석']]
                cls_sum = cls_full_days.groupby(['학년','반']).agg(체크인수=('날짜','count'),고유학생수=('이메일','nunique')).reset_index()
            else:
                cls_sum = df.groupby(['학년','반']).agg(체크인수=('이메일','count'),고유학생수=('이메일','nunique')).reset_index()
            cls_sum['반명'] = cls_sum.apply(lambda r: make_label(r['학년'],r['반']),axis=1)
            cls_valid = cls_sum[cls_sum['반명']!='미확인'].sort_values('체크인수', ascending=False).reset_index(drop=True)
            if cls_valid.empty:
                cls_top = cls_valid
            else:
                # ★ 6위와 동점인 반도 모두 "공동 6위"로 함께 표시
                cls_cutoff = cls_valid['체크인수'].iloc[min(5, len(cls_valid)-1)]
                cls_top = cls_valid[cls_valid['체크인수'] >= cls_cutoff].reset_index(drop=True)
            for chunk_start in range(0, len(cls_top), 3):
                chunk = cls_top.iloc[chunk_start:chunk_start+3]
                cols = st.columns(3)
                for col,(i,row) in zip(cols,chunk.iterrows()):
                    with col:
                        rank  = int((cls_top['체크인수'] > row['체크인수']).sum()) + 1
                        medal = MEDALS[rank-1] if rank<=3 else f"{rank}위"
                        cls   = ['gold','silver','bronze'][rank-1] if rank<=3 else ''
                        st.markdown(f"""<div class="top3-card {cls}">
                          <div style="font-size:1.5rem;text-align:center">{medal}</div>
                          <div style="text-align:center;font-size:1.1rem;font-weight:700">{row['반명']}</div>
                          <div style="text-align:center;color:#2d7ef7">완전출석 {int(row['체크인수'])}일</div>
                          <div style="text-align:center;color:#6b7280;font-size:0.85rem">참여 {int(row['고유학생수'])}명</div>
                        </div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════
# TAB 3: 주차별 추이
# ══════════════════════════════════════════════════
with tab3:
    start_date, end_date = period_filter_ui("t3")
    df = filter_period(df_valid, start_date, end_date)

    st.markdown("<div class='section-title'>📈 이번 주 vs 지난 주</div>", unsafe_allow_html=True)
    this_s,this_e = get_week_range(0); last_s,last_e = get_week_range(-1)
    df_this = filter_valid(filter_period(df_all,this_s,this_e))
    df_last = filter_valid(filter_period(df_all,last_s,last_e))
    this_cnt = df_this['이메일'].nunique() if not df_this.empty and '이메일' in df_this.columns else 0
    last_cnt = df_last['이메일'].nunique() if not df_last.empty and '이메일' in df_last.columns else 0

    c1,c2,c3,c4 = st.columns(4)
    with c1: st.metric("이번 주 학생",f"{this_cnt}명",delta=f"{this_cnt-last_cnt:+d}명")
    with c2: st.metric("이번 주 체크인",f"{len(df_this)}건")
    with c3: st.metric("지난 주 학생",f"{last_cnt}명")
    with c4: st.metric("지난 주 체크인",f"{len(df_last)}건")

    day_map = {0:'월',1:'화',2:'수',3:'목',4:'금'}
    def daily_cnt(d):
        if d.empty or '이메일' not in d.columns or '날짜' not in d.columns:
            return pd.DataFrame(columns=['요일번호','학생수','요일'])
        r = d.groupby(d['날짜'].dt.dayofweek)['이메일'].nunique().reset_index()
        r.columns = ['요일번호','학생수']; r['요일']=r['요일번호'].map(day_map)
        return r.sort_values('요일번호')

    fig = go.Figure()
    dc_this = daily_cnt(df_this); dc_last = daily_cnt(df_last)
    if not dc_this.empty:
        fig.add_trace(go.Scatter(x=dc_this['요일'],y=dc_this['학생수'],mode='lines+markers+text',
            name='이번 주',line=dict(color='#2d7ef7',width=3),marker=dict(size=10),
            text=dc_this['학생수'],textposition='top center'))
    if not dc_last.empty:
        fig.add_trace(go.Scatter(x=dc_last['요일'],y=dc_last['학생수'],mode='lines+markers+text',
            name='지난 주',line=dict(color='#94a3b8',width=2,dash='dash'),marker=dict(size=8),
            text=dc_last['학생수'],textposition='bottom center'))
    fig.update_layout(title="이번 주 vs 지난 주",plot_bgcolor='white',paper_bgcolor='white',
        legend=dict(orientation='h',y=1.1),height=340,margin=dict(t=50,b=20))
    fig.update_xaxes(gridcolor='#f3f4f6'); fig.update_yaxes(gridcolor='#f3f4f6')
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("<div class='section-title'>선택 기간 일별 추이</div>", unsafe_allow_html=True)
    if df.empty or '날짜' not in df.columns:
        empty_state("선택 기간에 데이터가 없습니다.")
    else:
        daily = df.groupby('날짜')['이메일'].nunique().reset_index().rename(columns={'이메일':'학생수'}).sort_values('날짜')
        daily['날짜str'] = daily['날짜'].dt.strftime('%m/%d(%a)')
        fig2 = px.area(daily,x='날짜str',y='학생수',title="일별 참여 학생 수",color_discrete_sequence=['#2d7ef7'])
        fig2.update_layout(plot_bgcolor='white',paper_bgcolor='white',height=280,margin=dict(t=40,b=20))
        st.plotly_chart(fig2, use_container_width=True)

# ══════════════════════════════════════════════════
# TAB 4: 학년·반별
# ══════════════════════════════════════════════════
with tab4:
    start_date, end_date = period_filter_ui("t4")
    df = filter_period(df_valid, start_date, end_date)
    st.markdown("<div class='section-title'>🏫 학년·반별 출석 현황</div>", unsafe_allow_html=True)

    if df.empty or '학년' not in df.columns:
        empty_state("선택 기간에 데이터가 없습니다.")
    else:
        valid_g = sorted([int(g) for g in df['학년'].dropna().unique()
                          if is_valid(g) and str(g).replace('.0','').isdigit()])
        sel_grade = st.selectbox("학년 선택",["전체"]+[f"{g}학년" for g in valid_g])
        df_g = df if sel_grade=="전체" else df[df['학년']==int(sel_grade.replace('학년',''))]

        if df_g.empty:
            empty_state(f"{sel_grade} 데이터 없음")
        elif '반' in df_g.columns and '이메일' in df_g.columns:
            sm = df_g.groupby(['학년','반']).agg(체크인수=('이메일','count'),고유학생수=('이메일','nunique')).reset_index()
            sm['반명'] = sm.apply(lambda r: make_label(r['학년'],r['반']),axis=1)
            sm = sm[sm['반명']!='미확인'].sort_values(['학년','반'])
            cl,cr = st.columns([3,2])
            with cl:
                fig = px.bar(sm,x='반명',y='고유학생수',color='학년',text='고유학생수',
                             title="반별 참여 학생 수",color_continuous_scale='Blues')
                fig.update_traces(textposition='outside')
                fig.update_layout(plot_bgcolor='white',paper_bgcolor='white',
                                  height=360,margin=dict(t=40,b=20),coloraxis_showscale=False)
                st.plotly_chart(fig, use_container_width=True)
            with cr:
                st.markdown("**반별 상세**")
                st.dataframe(sm[['반명','체크인수','고유학생수']].rename(
                    columns={'반명':'반','체크인수':'체크인','고유학생수':'학생수'}),
                    use_container_width=True, height=360)

# ══════════════════════════════════════════════════
# TAB 5: 어학과별
# ══════════════════════════════════════════════════
with tab5:
    start_date, end_date = period_filter_ui("t5")
    df = filter_period(df_valid, start_date, end_date)
    st.markdown("<div class='section-title'>🌍 어학과별 출석 현황</div>", unsafe_allow_html=True)

    if df.empty or '어학과' not in df.columns:
        empty_state("선택 기간에 데이터가 없습니다.")
    else:
        ds = df.groupby('어학과').agg(체크인수=('이메일','count'),고유학생수=('이메일','nunique'),
                                      운영일수=('날짜','nunique')).reset_index().sort_values('고유학생수',ascending=False)
        ds['1일평균'] = (ds['체크인수']/ds['운영일수'].replace(0,1)).round(1)
        cl,cr = st.columns(2)
        with cl:
            fig = px.pie(ds,names='어학과',values='고유학생수',title="어학과별 참여 비율",
                         color_discrete_sequence=px.colors.qualitative.Set2,hole=0.4)
            fig.update_traces(textinfo='label+percent',textfont_size=13)
            fig.update_layout(height=360,margin=dict(t=40,b=0),paper_bgcolor='white')
            st.plotly_chart(fig, use_container_width=True)
        with cr:
            fig2 = px.bar(ds,x='어학과',y='고유학생수',color='어학과',text='고유학생수',
                          title="어학과별 참여 학생 수",color_discrete_sequence=px.colors.qualitative.Set2)
            fig2.update_traces(textposition='outside',showlegend=False)
            fig2.update_layout(plot_bgcolor='white',paper_bgcolor='white',height=360,margin=dict(t=40,b=20))
            st.plotly_chart(fig2, use_container_width=True)
        if '날짜' in df.columns:
            dd = df.groupby(['날짜','어학과'])['이메일'].nunique().reset_index().rename(columns={'이메일':'학생수'})
            if not dd.empty:
                fig3 = px.line(dd,x='날짜',y='학생수',color='어학과',title="어학과별 일별 추이",
                               markers=True,color_discrete_sequence=px.colors.qualitative.Set2)
                fig3.update_layout(plot_bgcolor='white',paper_bgcolor='white',height=320,margin=dict(t=40,b=20))
                st.plotly_chart(fig3, use_container_width=True)
        st.dataframe(ds, use_container_width=True)

# ══════════════════════════════════════════════════
# TAB 6: 담임용 조회 — 학생목록 연동 + 요일별 출석표
# ══════════════════════════════════════════════════
with tab6:
    st.markdown("<div class='section-title'>👩‍🏫 담임용 우리 반 출석 조회</div>", unsafe_allow_html=True)

    valid_all_g = sorted([int(g) for g in df_valid['학년'].dropna().unique()
                          if is_valid(g) and str(g).replace('.0','').isdigit()]) \
                  if not df_valid.empty and '학년' in df_valid.columns else []

    if not valid_all_g:
        empty_state("유효한 학년 데이터가 없습니다.")
    else:
        if "t6_week_offset" not in st.session_state:
            st.session_state["t6_week_offset"] = 0

        hc1, hc2, hc3, hc4 = st.columns([1, 1, 2.5, 0.5])

        with hc1:
            sel_hg = st.selectbox("학년", [f"{g}학년" for g in valid_all_g], key='hgrade')
            h_gnum = int(sel_hg.replace('학년',''))

        with hc2:
            valid_hk = sorted([int(k) for k in df_valid[df_valid['학년']==h_gnum]['반'].dropna().unique()
                if is_valid(k) and str(k).replace('.0','').isdigit()]) if '반' in df_valid.columns else []
            if valid_hk:
                sel_hk = st.selectbox("반", [f"{k}반" for k in valid_hk], key='hklass')
                h_knum = int(sel_hk.replace('반',''))
            else:
                sel_hk = None; h_knum = None

        with hc3:
            ws, we = get_week_range(st.session_state["t6_week_offset"])
            week_label = f"{ws.month}/{ws.day}(월) ~ {we.month}/{we.day}(일)"
            st.markdown(f"<div style='font-size:12px;color:#6b7280;margin-bottom:6px'>조회 주간</div>", unsafe_allow_html=True)
            nav1, nav2, nav3 = st.columns([1,2,1])
            with nav1:
                if st.button("◀ 이전 주", use_container_width=True, key="t6_prev"):
                    st.session_state["t6_week_offset"] -= 1; st.rerun()
            with nav2:
                st.markdown(f"<div style='background:#f1f5f9;border-radius:8px;padding:7px 12px;text-align:center;font-size:13px;font-weight:600;color:#1d3a6e'>📅 {week_label}</div>", unsafe_allow_html=True)
            with nav3:
                if st.button("다음 주 ▶", use_container_width=True, key="t6_next"):
                    st.session_state["t6_week_offset"] += 1; st.rerun()

        with hc4:
            st.markdown("<div style='font-size:12px;color:#6b7280;margin-bottom:6px'>&nbsp;</div>", unsafe_allow_html=True)
            if st.button("🔄", use_container_width=True, key="t6_ref", help="새로고침"):
                st.cache_data.clear(); st.rerun()

        if not sel_hk or not h_knum:
            empty_state("반 데이터가 없습니다.")
        else:
            st.markdown("---")

            week_start = ws
            week_end   = we

            if not df_students.empty and '학년' in df_students.columns and '반' in df_students.columns:
                class_students = df_students[
                    (df_students['학년']==h_gnum) & (df_students['반']==h_knum)
                ].copy().sort_values('번호') if '번호' in df_students.columns else \
                df_students[(df_students['학년']==h_gnum) & (df_students['반']==h_knum)].copy()
            else:
                class_students = pd.DataFrame()

            df_week = df_valid[
                (df_valid['학년']==h_gnum) &
                (df_valid['반']==h_knum) &
                (df_valid['날짜'] >= pd.Timestamp(week_start)) &
                (df_valid['날짜'] <= pd.Timestamp(week_end))
            ].copy() if not df_valid.empty else pd.DataFrame()

            total_students    = len(class_students) if not class_students.empty else \
                                (df_week['이름'].nunique() if not df_week.empty and '이름' in df_week.columns else 0)
            attended_students = df_week['이름'].nunique() if not df_week.empty and '이름' in df_week.columns else 0
            not_attended      = max(0, total_students - attended_students)
            total_checkins    = len(df_week)

            m1,m2,m3,m4 = st.columns(4)
            with m1: st.metric("전체 학생", f"{total_students}명")
            with m2: st.metric("이번 주 참여", f"{attended_students}명")
            with m3: st.metric("미참여", f"{not_attended}명",
                               delta=f"-{not_attended}" if not_attended>0 else "0",
                               delta_color="inverse")
            with m4: st.metric("총 체크인", f"{total_checkins}건")

            st.markdown("---")

            dow_kor  = {0:'월',1:'화',2:'수',3:'목',4:'금'}
            week_days = []
            d = week_start
            while d <= week_end:
                if d.weekday() < 5:
                    week_days.append(d)
                d += timedelta(days=1)

            if not class_students.empty and '이름' in class_students.columns:
                if '번호' in class_students.columns:
                    student_rows = class_students[['번호','이름']].drop_duplicates().values.tolist()
                else:
                    student_rows = [[None, n] for n in class_students['이름'].unique()]
            elif not df_week.empty and '이름' in df_week.columns:
                if '번호' in df_week.columns:
                    student_rows = df_week[['번호','이름']].drop_duplicates().sort_values('번호').values.tolist()
                else:
                    student_rows = [[None, n] for n in df_week['이름'].unique()]
            else:
                student_rows = []

            # ★ 범례: '2~3교시' → '3교시'
            st.markdown(f"""
            <div style='display:flex;align-items:center;justify-content:space-between;margin-bottom:10px'>
              <div style='font-size:1.05rem;font-weight:700;color:#1d3a6e'>
                📋 {h_gnum}학년 {h_knum}반 — 학생별 요일·교시 출석 현황
              </div>
              <div style='display:flex;gap:8px;font-size:12px;align-items:center;flex-wrap:wrap'>
                <span style='background:#dbeafe;color:#1e40af;padding:2px 10px;border-radius:4px;font-weight:600'>1교시</span>
                <span style='background:#ede9fe;color:#5b21b6;padding:2px 10px;border-radius:4px;font-weight:600'>3교시</span>
                <span style='background:#d1fae5;color:#065f46;padding:2px 10px;border-radius:4px;font-weight:600'>둘 다</span>
                <span style='color:#9ca3af'>— 미참여</span>
              </div>
            </div>""", unsafe_allow_html=True)

            if not student_rows:
                empty_state("표시할 학생 데이터가 없습니다.")
            else:
                day_headers = ""
                for day in week_days:
                    dow_str  = dow_kor[day.weekday()]
                    date_str = f"{day.month}/{day.day}"
                    day_headers += f"<th style='padding:10px 6px;text-align:center;min-width:76px'>{dow_str}<br><span style='font-size:11px;opacity:0.8;font-weight:400'>{date_str}</span></th>"

                tbody_rows = ""
                for s_info in student_rows:
                    s_num  = s_info[0] if len(s_info) > 0 else None
                    s_name = s_info[1] if len(s_info) > 1 else ""

                    week_total = 0
                    row_cells  = ""

                    for day in week_days:
                        day_ts = pd.Timestamp(day)
                        if not df_week.empty and '날짜' in df_week.columns and '이름' in df_week.columns:
                            day_rec = df_week[(df_week['날짜']==day_ts) & (df_week['이름']==s_name)]
                            periods = set(day_rec['교시'].tolist()) if '교시' in day_rec.columns and not day_rec.empty else set()
                        else:
                            periods = set()

                        has_p1 = '1교시' in periods
                        # ★ '2~3교시' → '3교시' 로 통일
                        has_p2 = '3교시' in periods

                        if has_p1 and has_p2:
                            cell = "<span style='background:#d1fae5;color:#065f46;font-size:11px;padding:2px 8px;border-radius:4px;font-weight:600'>둘 다</span>"
                            week_total += 2
                        elif has_p1:
                            cell = "<span style='background:#dbeafe;color:#1e40af;font-size:11px;padding:2px 8px;border-radius:4px;font-weight:600'>1교시</span>"
                            week_total += 1
                        elif has_p2:
                            # ★ 셀 표시도 '3교시'
                            cell = "<span style='background:#ede9fe;color:#5b21b6;font-size:11px;padding:2px 8px;border-radius:4px;font-weight:600'>3교시</span>"
                            week_total += 1
                        else:
                            cell = "<span style='color:#d1d5db;font-size:14px'>—</span>"

                        row_cells += f"<td style='padding:10px 6px;text-align:center;border-bottom:1px solid #f3f4f6'>{cell}</td>"

                    if week_total == 0:
                        badge = f"<span style='background:#fef2f2;color:#991b1b;border:1px solid #fecaca;border-radius:12px;padding:2px 10px;font-size:11px;font-weight:600'>0회</span>"
                    elif week_total >= 8:
                        badge = f"<span style='background:#d1fae5;color:#065f46;border:1px solid #6ee7b7;border-radius:12px;padding:2px 10px;font-size:11px;font-weight:600'>{week_total}회 ⭐</span>"
                    else:
                        badge = f"<span style='background:#f0fdf4;color:#166534;border:1px solid #bbf7d0;border-radius:12px;padding:2px 10px;font-size:11px;font-weight:600'>{week_total}회</span>"

                    num_str = f"<span style='font-size:10px;color:#9ca3af;display:block'>{int(s_num)}번</span>" if s_num else ""
                    row_bg  = "#fef2f2" if week_total == 0 else "white"

                    tbody_rows += f"""<tr style='background:{row_bg}'>
                      <td style='padding:10px 14px;border-bottom:1px solid #f3f4f6;white-space:nowrap'>
                        {num_str}<span style='font-weight:600'>{s_name}</span>
                      </td>
                      {row_cells}
                      <td style='padding:10px 6px;text-align:center;border-bottom:1px solid #f3f4f6'>{badge}</td>
                    </tr>"""

                st.markdown(f"""
                <div style='overflow-x:auto;border-radius:16px;box-shadow:0 4px 24px rgba(26,47,90,0.08);margin-bottom:12px'>
                <table style='width:100%;border-collapse:collapse;font-size:13px'>
                  <thead>
                    <tr style='background:#1d3a6e;color:white'>
                      <th style='padding:12px 14px;text-align:left;min-width:90px'>학생</th>
                      {day_headers}
                      <th style='padding:10px 6px;text-align:center;min-width:72px'>주간 합계</th>
                    </tr>
                  </thead>
                  <tbody>{tbody_rows}</tbody>
                </table>
                </div>
                <div style='font-size:11px;color:#9ca3af;margin-bottom:16px'>
                  ⭐ 주간 8회 이상(1+3교시 모두) &nbsp;·&nbsp; 🔴 미참여 학생은 행 배경 연빨간색 &nbsp;·&nbsp; 학생목록 시트 기준
                </div>""", unsafe_allow_html=True)

            st.markdown("<div class='section-title'>📈 이번 주 일별 출석 추이</div>", unsafe_allow_html=True)
            if not df_week.empty and '날짜' in df_week.columns and '이메일' in df_week.columns:
                dc = df_week.groupby('날짜')['이메일'].nunique().reset_index().rename(columns={'이메일':'학생수'})
                dc['날짜str'] = dc['날짜'].dt.strftime('%m/%d(%a)')
                figc = px.bar(dc, x='날짜str', y='학생수',
                              title=f"{h_gnum}학년 {h_knum}반 일별 출석",
                              color_discrete_sequence=['#2d7ef7'], text='학생수')
                figc.update_traces(textposition='outside')
                figc.update_layout(plot_bgcolor='white', paper_bgcolor='white',
                                   height=260, margin=dict(t=40,b=20))
                st.plotly_chart(figc, use_container_width=True)
            else:
                empty_state("이번 주 출석 데이터가 없습니다.")

# ── 학생 개인 검색 ────────────────────────────────────
st.markdown("---")
st.markdown("<div class='section-title'>🔍 학생 개인 검색</div>", unsafe_allow_html=True)
search = st.text_input("이름 또는 이메일로 검색", placeholder="예: 홍길동 또는 student@hyfl.hs.kr")
if search and not df_valid.empty:
    mask = pd.Series([False]*len(df_valid), index=df_valid.index)
    if '이름'  in df_valid.columns: mask = mask | df_valid['이름'].astype(str).str.contains(search,na=False)
    if '이메일' in df_valid.columns: mask = mask | df_valid['이메일'].astype(str).str.contains(search,na=False)
    ds = df_valid[mask]
    if ds.empty:
        empty_state(f"'{search}' 검색 결과가 없습니다.")
    else:
        s1,s2,s3 = st.columns(3)
        with s1: st.metric("총 체크인",f"{len(ds)}건")
        with s2: st.metric("출석 일수",f"{ds['날짜'].nunique() if '날짜' in ds.columns else 0}일")
        with s3: st.metric("교시 종류",f"{ds['교시'].nunique() if '교시' in ds.columns else 0}종류")
        sc = [c for c in ['날짜','교시','좌석','학년','반','번호','이름','시간'] if c in ds.columns]
        st.dataframe(ds[sc].sort_values('날짜',ascending=False) if '날짜' in sc else ds[sc],
                     use_container_width=True, height=300)
